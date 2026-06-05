"""Services de génération des documents de sortie (Sprint 5).

Chaque fonction renvoie les octets du fichier généré (bytes), prêts à être
servis dans une réponse HTTP. Aucune écriture sur disque.

  - build_participants_xlsx   → EXP-01 (Excel)
  - build_participant_cni_pdf → EXP-02 (PDF fiche CNI individuelle)
  - build_presence_list_pdf   → EXP-03 (PDF liste de présence à signer)
  - build_cni_zip             → EXP-04 (ZIP des fiches CNI)
  - build_activite_qr_pdf     → Affiche du QR Code (infos activité + QR)
"""

import io
import zipfile

import qrcode
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Palette alignée sur le thème de l'interface (cyan).
_CYAN = "0891B2"
_CYAN_DARK = "0E7490"
_ZEBRA = "F0F9FB"
_INK = "0F172A"


def _fmt(dt) -> str:
    """Formate une date en heure locale, tolérant les valeurs nulles."""
    if not dt:
        return ""
    return timezone.localtime(dt).strftime("%d/%m/%Y %H:%M")


def _fmt_date(dt) -> str:
    """Formate une date (sans l'heure)."""
    if not dt:
        return ""
    return timezone.localtime(dt).strftime("%d/%m/%Y")


def _safe(name: str) -> str:
    """Normalise une chaîne pour un nom de fichier (pas de caractères spéciaux)."""
    keep = "".join(c if c.isalnum() or c in " -_" else "_" for c in name)
    return "_".join(keep.split()).strip("_") or "export"


# --- EXP-01 : Export Excel -------------------------------------------------


def build_participants_xlsx(activite, participants) -> bytes:
    """Classeur .xlsx : onglet récapitulatif + liste formatée des participants."""
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor=_CYAN)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color=_INK, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D5DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    colonnes = [
        "N°",
        "Nom",
        "Prénom",
        "Structure",
        "Fonction",
        "Téléphone Wave",
        "Email",
        "N° CNI",
        "Horodatage",
    ]
    nb_cols = len(colonnes)
    derniere_col = get_column_letter(nb_cols)

    ws = wb.active
    ws.title = "Participants"

    # --- Bloc récapitulatif (en haut de l'unique feuille) ------------------
    complets = sum(1 for p in participants if p.photo_cni_recto and p.photo_cni_verso)
    recap_font = Font(color="475569", size=10)

    ws.merge_cells(f"A1:{derniere_col}1")
    ws["A1"] = activite.nom
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    recap_lignes = [
        f"Ville : {activite.ville}     •     "
        f"Lieu : {activite.lieu}     •     "
        f"Période : {_fmt(activite.date_debut)} → {_fmt(activite.date_fin)}     •     "
        f"Statut : {activite.get_statut_display()}     •     "
        f"Organisateur : {activite.created_by.username}",
        f"Participants : {len(participants)}     •     "
        f"CNI complètes : {complets}     •     "
        f"CNI incomplètes : {len(participants) - complets}     •     "
        f"Document généré le : {_fmt(timezone.now())}",
    ]
    for i, texte in enumerate(recap_lignes, start=2):
        ws.merge_cells(f"A{i}:{derniere_col}{i}")
        ws[f"A{i}"] = texte
        ws[f"A{i}"].font = recap_font
        ws[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Tableau des participants (sous le récapitulatif) ------------------
    header_row = len(recap_lignes) + 3  # 1 ligne titre + N lignes récap + 1 vide
    max_len = [len(titre) for titre in colonnes]

    for col, titre in enumerate(colonnes, start=1):
        cell = ws.cell(row=header_row, column=col, value=titre)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for idx, p in enumerate(participants, start=1):
        row = header_row + idx
        valeurs = [
            idx,
            p.nom,
            p.prenom,
            p.structure,
            p.fonction,
            p.telephone_wave,
            p.email,
            p.numero_cni,
            _fmt(p.horodatage),
        ]
        for col, value in enumerate(valeurs, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center if col in (1, 6, 8, 9) else left
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            max_len[col - 1] = max(max_len[col - 1], len(str(value)))

    # Auto-dimensionnement : largeur = contenu le plus long + marge, bornée.
    for col, longueur in enumerate(max_len, start=1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(longueur + 3, 8), 50)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = (
        f"A{header_row}:{derniere_col}{header_row + len(participants)}"
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- EXP-02 : PDF fiche CNI individuelle -----------------------------------


def build_participant_cni_pdf(activite, participant) -> bytes:
    """Fiche PDF : en-tête, recto/verso empilés à taille réelle, pied de page."""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    footer_h = 16 * mm

    # Taille réelle d'une carte (format carte bancaire : 85,6 x 54 mm),
    # empilées verticalement et centrées entre le haut de page et le pied.
    img_w = 85.6 * mm
    img_h = 54 * mm
    gap = 14 * mm
    total_h = 2 * img_h + gap
    x = (width - img_w) / 2
    zone_bas = footer_h
    zone_haut = height - footer_h
    bas_bloc = zone_bas + (zone_haut - zone_bas - total_h) / 2
    y_verso = bas_bloc
    y_recto = bas_bloc + img_h + gap

    for field, y in (
        (participant.photo_cni_recto, y_recto),
        (participant.photo_cni_verso, y_verso),
    ):
        c.setStrokeColor(colors.HexColor("#CBD5E1"))
        c.rect(x, y, img_w, img_h, fill=0, stroke=1)
        if field:
            try:
                c.drawImage(
                    field.path,
                    x,
                    y,
                    width=img_w,
                    height=img_h,
                    preserveAspectRatio=True,
                    anchor="c",
                )
            except Exception:
                _placeholder(c, x, y, img_w, img_h, "Image illisible")
        else:
            _placeholder(c, x, y, img_w, img_h, "Photo non fournie")

    # Pied de page
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.line(18 * mm, footer_h, width - 18 * mm, footer_h)
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.drawString(18 * mm, 9 * mm, f"{activite.nom}  ·  Fiche générée le {_fmt(timezone.now())}")
    c.drawRightString(width - 18 * mm, 9 * mm, "Page 1 / 1")

    c.showPage()
    c.save()
    return buffer.getvalue()


def _placeholder(c, x, y, w, h, text: str) -> None:
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica-Oblique", 9)
    c.drawCentredString(x + w / 2, y + h / 2, text)
    c.setFillColor(colors.HexColor(f"#{_INK}"))


# --- EXP-03 : PDF liste de présence à signer -------------------------------


def build_presence_list_pdf(activite, participants) -> bytes:
    """Liste de présence A4 paysage avec une colonne Signature vide."""
    page = landscape(A4)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Liste de présence — {activite.nom}",
    )
    styles = getSampleStyleSheet()
    titre = styles["Title"]
    titre.fontSize = 16
    sous = styles["Normal"]
    sous.fontSize = 9
    cell = styles["Normal"].clone("cell")
    cell.fontSize = 8

    elements = [
        Paragraph(f"Liste de présence — {activite.nom}", titre),
        Paragraph(
            f"{activite.ville} &nbsp;·&nbsp; {activite.lieu} &nbsp;·&nbsp; "
            f"{_fmt(activite.date_debut)} → {_fmt(activite.date_fin)}",
            sous,
        ),
        Spacer(1, 8 * mm),
    ]

    entetes = [
        "N°",
        "Nom",
        "Prénom",
        "Structure",
        "Fonction",
        "Téléphone",
        "Signature",
    ]
    data = [entetes]
    for i, p in enumerate(participants, start=1):
        data.append(
            [
                str(i),
                Paragraph(p.nom, cell),
                Paragraph(p.prenom, cell),
                Paragraph(p.structure, cell),
                Paragraph(p.fonction, cell),
                p.telephone_wave,
                "",
            ]
        )
    if not participants:
        data.append(["—", "Aucun participant", "", "", "", "", ""])

    # Largeurs étalées sur la largeur paysage (~267 mm utiles).
    table = Table(
        data,
        colWidths=[
            12 * mm,
            38 * mm,
            34 * mm,
            50 * mm,
            42 * mm,
            34 * mm,
            57 * mm,
        ],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_CYAN}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (5, 0), (5, -1), "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{_ZEBRA}")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ]
        )
    )
    elements.append(table)

    def _footer(c, _doc):
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#94A3B8"))
        c.drawString(15 * mm, 10 * mm, f"Généré le {_fmt(timezone.now())}")
        c.drawRightString(page[0] - 15 * mm, 10 * mm, f"Page {c.getPageNumber()}")

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


# --- EXP-04 : ZIP des fiches CNI -------------------------------------------


def build_cni_zip(activite, participants) -> bytes:
    """Archive ZIP regroupant les fiches PDF CNI de chaque participant."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        used: set[str] = set()
        for p in participants:
            base = f"{_safe(p.nom)}_{_safe(p.prenom)}_cni"
            name = f"{base}.pdf"
            n = 2
            while name in used:
                name = f"{base}_{n}.pdf"
                n += 1
            used.add(name)
            zf.writestr(name, build_participant_cni_pdf(activite, p))
    return buffer.getvalue()


# --- Affiche du QR Code (infos activité + QR) ------------------------------


def build_activite_qr_pdf(activite) -> bytes:
    """Affiche A4 : informations de l'activité (date sans l'heure) + QR Code."""
    url = f"{settings.PUBLIC_FORM_BASE_URL}/form/{activite.token_qr}"
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Bandeau d'accent
    c.setFillColor(colors.HexColor(f"#{_CYAN_DARK}"))
    c.rect(0, height - 10 * mm, width, 10 * mm, fill=1, stroke=0)

    # Titre
    c.setFillColor(colors.HexColor(f"#{_INK}"))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(width / 2, height - 32 * mm, activite.nom[:55])

    # Informations (date sans l'heure)
    d1, d2 = _fmt_date(activite.date_debut), _fmt_date(activite.date_fin)
    periode = d1 if d1 == d2 else f"{d1} au {d2}"
    infos = [
        f"Ville : {activite.ville}",
        f"Lieu : {activite.lieu}",
        f"Date : {periode}",
    ]
    c.setFont("Helvetica", 13)
    c.setFillColor(colors.HexColor("#475569"))
    y = height - 46 * mm
    for ligne in infos:
        c.drawCentredString(width / 2, y, ligne)
        y -= 8 * mm

    # QR Code (centré, grand)
    qr_img = qrcode.make(url)
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_buf.seek(0)
    size = 95 * mm
    qr_y = (height - size) / 2 - 18 * mm
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.rect((width - size) / 2 - 6 * mm, qr_y - 6 * mm, size + 12 * mm, size + 12 * mm)
    c.drawImage(ImageReader(qr_buf), (width - size) / 2, qr_y, size, size)

    # Instruction + URL
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(colors.HexColor(f"#{_CYAN_DARK}"))
    c.drawCentredString(
        width / 2, qr_y - 16 * mm, "Scannez ce QR Code pour vous enregistrer"
    )
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.drawCentredString(width / 2, qr_y - 24 * mm, url)

    c.showPage()
    c.save()
    return buffer.getvalue()
